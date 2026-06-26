#!/usr/bin/env python3
"""
Gerador de Ordem de Serviço (OS)
=================================
Abre planilha Excel, seleciona patrimônio e periodicidade,
gera OS na aba OSP e registra no Movimento.

Uso:
    python main.py
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import xlrd
import openpyxl
import os
import math
from datetime import datetime, timedelta
import traceback

# =====================================================================
# UTILITÁRIOS
# =====================================================================

def safe_str(val):
    """Converte qualquer valor para string de forma segura."""
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    return str(val).strip()

def safe_float(val):
    """Converte para float, retornando 0.0 se falhar."""
    if val is None:
        return 0.0
    if isinstance(val, float) and math.isnan(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip().replace(",", ".")
        try:
            return float(val)
        except ValueError:
            return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def is_numeric(val):
    """Verifica se o valor pode ser interpretado como número."""
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return not (isinstance(val, float) and math.isnan(val))
    if isinstance(val, str):
        val = val.strip().replace(",", ".")
        try:
            float(val)
            return True
        except ValueError:
            return False
    return False

def excel_serial_to_date(serial, datemode=0):
    """Converte número serial do Excel para datetime."""
    if serial is None:
        return None
    try:
        serial = float(serial)
    except (ValueError, TypeError):
        return None
    if serial < 1:
        return None
    try:
        return xlrd.xldate_as_datetime(serial, datemode)
    except Exception:
        base = datetime(1899, 12, 30)
        return base + timedelta(days=serial)


# =====================================================================
# MACRO ENGINE — Apenas geração de OS
# =====================================================================

class MacroEngine:
    """Motor que manipula a planilha para geração de OS."""

    def __init__(self):
        self.workbook = None
        self.filepath = None
        self._datemode = 0

    # ------------------------------------------------------------------
    # CARGA / SALVAMENTO
    # ------------------------------------------------------------------

    def load_xls(self, filepath):
        """Carrega um arquivo .xls e converte para openpyxl Workbook."""
        self.filepath = filepath
        wb_old = xlrd.open_workbook(filepath)
        self._datemode = wb_old.datemode

        wb_new = openpyxl.Workbook()
        wb_new.remove(wb_new.active)

        for sh_name in wb_old.sheet_names():
            sh_old = wb_old.sheet_by_name(sh_name)
            sh_new = wb_new.create_sheet(title=sh_name[:31])

            for r in range(sh_old.nrows):
                for c in range(sh_old.ncols):
                    cell_old = sh_old.cell(r, c)
                    cell_new = sh_new.cell(row=r + 1, column=c + 1)
                    ctype = cell_old.ctype
                    value = cell_old.value

                    if ctype == 0:
                        continue
                    elif ctype == 1:
                        cell_new.value = value
                    elif ctype == 2:
                        cell_new.value = value
                    elif ctype == 3:
                        try:
                            if hasattr(value, '__iter__') and not isinstance(value, str):
                                parts = [int(v) if v is not None else 0 for v in value[:6]]
                                while len(parts) < 6:
                                    parts.append(0)
                                dt = datetime(*parts)
                            else:
                                dt = excel_serial_to_date(value)
                            if dt:
                                cell_new.value = dt
                            else:
                                cell_new.value = str(value)
                        except Exception:
                            cell_new.value = str(value)
                    elif ctype == 4:
                        cell_new.value = bool(value)
                    elif ctype == 5:
                        cell_new.value = None
                    else:
                        continue

        self.workbook = wb_new
        return wb_new

    def load_xlsx(self, filepath):
        """Carrega um arquivo .xlsx diretamente (preserva formatação)."""
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath)
        self._datemode = 0
        return self.workbook

    def save(self, filepath=None):
        """Salva o workbook no formato .xlsx."""
        if filepath:
            self.filepath = filepath
        if not self.filepath:
            return False
        self.workbook.save(self.filepath)
        return True

    def get_sheet_names(self):
        """Retorna lista de nomes de abas."""
        if self.workbook:
            return self.workbook.sheetnames
        return []

    def get_sheet_data(self, sheet_name, max_cols=50):
        """Retorna dados de uma aba como lista de listas."""
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []
        ws = self.workbook[sheet_name]
        max_row = ws.max_row or 0
        if max_row == 0:
            return []

        actual_cols = 0
        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 100),
                                min_col=1, max_col=min(max_cols, 50), values_only=True):
            for c, v in enumerate(row):
                if v is not None and (not isinstance(v, float) or not math.isnan(v)):
                    if c + 1 > actual_cols:
                        actual_cols = c + 1

        actual_cols = max(actual_cols, 1)
        actual_cols = min(actual_cols, max_cols)

        data = []
        for row in ws.iter_rows(min_row=1, max_row=max_row,
                                min_col=1, max_col=actual_cols, values_only=False):
            row_data = []
            all_none = True
            for cell in row:
                v = cell.value
                if v is not None and (not isinstance(v, float) or not math.isnan(v)):
                    all_none = False
                row_data.append(v)
            if not all_none:
                data.append(row_data)

        return data

    # ------------------------------------------------------------------
    # PREENCHE OSP
    # ------------------------------------------------------------------

    def preenche_osp(self, progress_callback=None):
        """
        Implementa PreencheOSP() do VBA:
        - Lê dados do cabeçalho da OSP
        - Busca serviços em ParametroManut
        - Popula lista de serviços na OSP
        - Copia registro para Movimento
        - Incrementa número da OS
        """
        if not self.workbook:
            raise ValueError("Nenhum arquivo carregado.")

        needed = {"OSP", "Movimento", "ParametroManut"}
        missing = needed - set(self.workbook.sheetnames)
        if missing:
            raise ValueError(f"Abas necessárias não encontradas: {', '.join(missing)}")

        ws_osp = self.workbook["OSP"]
        ws_mov = self.workbook["Movimento"]
        ws_param = self.workbook["ParametroManut"]

        # --- Lê cabeçalho da OSP ---
        os_num = ws_osp.cell(row=1, column=9).value
        if os_num is None:
            os_num = 1
        try:
            os_num = int(float(str(os_num).strip()))
        except (ValueError, TypeError):
            os_num = 1

        a6 = safe_str(ws_osp.cell(row=6, column=1).value)   # Patrimônio
        c4 = safe_str(ws_osp.cell(row=4, column=3).value)   # Equipamento
        f4 = safe_str(ws_osp.cell(row=4, column=6).value)   # Marca
        h4 = safe_str(ws_osp.cell(row=4, column=8).value)   # Modelo
        a4 = safe_str(ws_osp.cell(row=4, column=1).value)   # Local

        b6_val = ws_osp.cell(row=6, column=2).value          # TipoManut
        c6_val = ws_osp.cell(row=6, column=3).value          # Periodicidade

        b6_str = safe_str(b6_val)
        if is_numeric(c6_val):
            c6_str = str(int(safe_float(c6_val)))
        else:
            c6_str = safe_str(c6_val)
        lookup_key = b6_str + c6_str

        d6 = safe_str(ws_osp.cell(row=6, column=4).value)   # Unde
        e6 = ws_osp.cell(row=6, column=5).value              # Ult.Revisão
        f6 = ws_osp.cell(row=6, column=6).value              # Prox.Manut.
        h6 = ws_osp.cell(row=6, column=8).value              # Emissão

        if progress_callback:
            progress_callback(f"OS nº {os_num} — Equipamento: {c4}")
            progress_callback(f"Patrimônio: {a6} | Tipo: {b6_str} | Per.: {c6_str}")

        # --- Armazena chave na primeira linha disponível do ParametroManut ---
        ult_col = 1
        for col in range(1, (ws_param.max_column or 1) + 2):
            val = ws_param.cell(row=1, column=col).value
            if val is not None and (not isinstance(val, float) or not math.isnan(val)):
                ult_col = col + 1
        ws_param.cell(row=1, column=ult_col).value = lookup_key

        # --- Limpa área de serviços A9:H50 ---
        # Cada linha A9:H9..A50:H50 é célula mesclada, basta limpar col A
        for r in range(9, 51):
            ws_osp.cell(row=r, column=1).value = None

        if progress_callback:
            progress_callback(f"Chave de busca: '{lookup_key}'")

        # --- Localiza a coluna do serviço no ParametroManut ---
        coluna_proc = None
        for col in range(1, (ws_param.max_column or 1) + 1):
            val = ws_param.cell(row=1, column=col).value
            if val is not None and str(val).strip() == lookup_key:
                coluna_proc = col
                break

        if coluna_proc is None:
            coluna_proc = ult_col

        if coluna_proc == ult_col and progress_callback:
            progress_callback("⚠ Parâmetro do equipamento não consta na lista de serviços!")

        # --- Coleta serviços marcados com "X" ---
        servicos = []
        for r in range(2, (ws_param.max_row or 1) + 1):
            val = ws_param.cell(row=r, column=coluna_proc).value
            if val is not None and safe_str(val).upper() == "X":
                desc = safe_str(ws_param.cell(row=r, column=1).value)
                servicos.append(desc)

        if progress_callback:
            progress_callback(f"{len(servicos)} serviços encontrados.")

        # --- Popula OSP ---
        for i, desc in enumerate(servicos):
            if i < 42:
                ws_osp.cell(row=9 + i, column=1).value = desc

        # --- Limpa chave temporária ---
        ws_param.cell(row=1, column=ult_col).value = None

        # --- Copia para Movimento ---
        ult_lin_mov = (ws_mov.max_row or 1) + 1
        for r in range((ws_mov.max_row or 1), 0, -1):
            val = ws_mov.cell(row=r, column=1).value
            if val is not None and (not isinstance(val, float) or not math.isnan(val)):
                ult_lin_mov = r + 1
                break

        row_mov = ult_lin_mov
        ws_mov.cell(row=row_mov, column=1).value = os_num       # OSP
        ws_mov.cell(row=row_mov, column=2).value = a6           # Cod. Patrim
        ws_mov.cell(row=row_mov, column=3).value = c4           # Bem
        ws_mov.cell(row=row_mov, column=4).value = f4           # Marca
        ws_mov.cell(row=row_mov, column=5).value = h4           # Modelo
        ws_mov.cell(row=row_mov, column=6).value = a4           # Localizacao
        ws_mov.cell(row=row_mov, column=7).value = b6_str       # TipoManut
        ws_mov.cell(row=row_mov, column=8).value = c6_str       # Periodicidade
        ws_mov.cell(row=row_mov, column=9).value = d6           # Unde
        ws_mov.cell(row=row_mov, column=10).value = e6          # Ult.Manut.
        ws_mov.cell(row=row_mov, column=11).value = f6          # Prox.Manut.
        ws_mov.cell(row=row_mov, column=12).value = h6          # Emissão

        if progress_callback:
            progress_callback(f"Registro copiado para Movimento (linha {row_mov}).")

        # --- Incrementa número da OS ---
        novo_num = os_num + 1
        ws_osp.cell(row=1, column=9).value = novo_num
        if progress_callback:
            progress_callback(f"OS nº {os_num} finalizada. Próximo nº: {novo_num}.")
            progress_callback("PreencheOSP concluído com sucesso!")

        return True


# =====================================================================
# DIÁLOGO DE GERAÇÃO DE OS
# =====================================================================

class GerarOSDialog:
    """Diálogo modal para geração assistida de Ordem de Serviço."""

    def __init__(self, parent, engine):
        self.parent = parent
        self.engine = engine

        self.dialog = tk.Toplevel(parent.root)
        self.dialog.title("Gerar Ordem de Serviço")
        self.dialog.geometry("720x620")
        self.dialog.minsize(600, 500)
        self.dialog.transient(parent.root)
        self.dialog.grab_set()

        self.patrimonio_var = tk.StringVar()
        self.descricao_var = tk.StringVar()
        self.marca_var = tk.StringVar()
        self.modelo_var = tk.StringVar()
        self.local_var = tk.StringVar()
        self.tipo_manut_var = tk.StringVar()
        self.unde_var = tk.StringVar()
        self.periodicidade_var = tk.StringVar()

        self.cadastro_rows = []
        self._load_cadastro()
        self._build_ui()

        self.dialog.update_idletasks()
        px = parent.root.winfo_x() + (parent.root.winfo_width() - self.dialog.winfo_width()) // 2
        py = parent.root.winfo_y() + (parent.root.winfo_height() - self.dialog.winfo_height()) // 2
        self.dialog.geometry(f"+{max(0, px)}+{max(0, py)}")

    def _load_cadastro(self):
        """Carrega dados da aba CadastroPat."""
        if "CadastroPat" not in self.engine.workbook.sheetnames:
            raise ValueError("Aba 'CadastroPat' não encontrada na planilha.")
        ws = self.engine.workbook["CadastroPat"]
        self.cadastro_rows = []
        for r in range(2, (ws.max_row or 1) + 1):
            row = {
                'patrimonio': safe_str(ws.cell(row=r, column=1).value),
                'descricao': safe_str(ws.cell(row=r, column=2).value),
                'marca': safe_str(ws.cell(row=r, column=3).value),
                'modelo': safe_str(ws.cell(row=r, column=4).value),
                'local': safe_str(ws.cell(row=r, column=5).value),
                'tipo_manut': safe_str(ws.cell(row=r, column=6).value),
                'unde': safe_str(ws.cell(row=r, column=7).value),
            }
            if row['patrimonio']:
                self.cadastro_rows.append(row)

    def _build_ui(self):
        """Constrói o layout completo do diálogo."""
        main = ttk.Frame(self.dialog, padding=14)
        main.pack(fill=tk.BOTH, expand=True)

        # --- Patrimônio ---
        sel_frame = ttk.LabelFrame(main, text=" Seleção de Patrimônio ", padding=8)
        sel_frame.pack(fill=tk.X)

        ttk.Label(sel_frame, text="Patrimônio:").grid(row=0, column=0, sticky=tk.W, padx=(0, 6))
        patri_list = [r['patrimonio'] for r in self.cadastro_rows]
        self.patrimonio_combo = ttk.Combobox(sel_frame, textvariable=self.patrimonio_var,
                                              values=patri_list, width=30, state="normal")
        self.patrimonio_combo.grid(row=0, column=1, sticky=tk.EW, padx=(0, 6))
        self.patrimonio_combo.bind("<<ComboboxSelected>>", self._on_patrimonio_select)
        sel_frame.columnconfigure(1, weight=1)

        ttk.Button(sel_frame, text="Buscar", width=8,
                   command=self._on_patrimonio_select).grid(row=0, column=2)

        # --- Dados do equipamento ---
        eq_frame = ttk.LabelFrame(main, text=" Dados do Equipamento ", padding=8)
        eq_frame.pack(fill=tk.X, pady=(8, 0))

        labels = [
            ("Tipo Manutenção:", 0, 0, self.tipo_manut_var),
            ("Descrição:", 1, 0, self.descricao_var),
            ("Marca:", 0, 3, self.marca_var),
            ("Modelo:", 1, 3, self.modelo_var),
            ("Localização:", 0, 6, self.local_var),
            ("Unde:", 1, 6, self.unde_var),
        ]
        for text, row, col, var in labels:
            ttk.Label(eq_frame, text=text).grid(row=row, column=col, sticky=tk.W, padx=(0, 4), pady=2)
            entry = ttk.Entry(eq_frame, textvariable=var, width=22, state="readonly")
            entry.grid(row=row, column=col + 1, sticky=tk.W, padx=(0, 16), pady=2)

        # --- Periodicidade ---
        per_frame = ttk.Frame(main)
        per_frame.pack(fill=tk.X, pady=(8, 0))
        ttk.Label(per_frame, text="Periodicidade (KM/Horas):").pack(side=tk.LEFT, padx=(0, 6))
        self.per_combo = ttk.Combobox(per_frame, textvariable=self.periodicidade_var,
                                       values=["5000", "10000", "15000", "20000",
                                               "25000", "30000", "40000", "50000"],
                                       width=14, state="normal")
        self.per_combo.pack(side=tk.LEFT)
        self.per_combo.bind("<<ComboboxSelected>>", self._atualizar_preview)
        self.periodicidade_var.trace_add("write", lambda *_: self._atualizar_preview())

        # --- Preview de serviços ---
        preview_frame = ttk.LabelFrame(main, text=" Serviços a Executar ", padding=8)
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 0))

        cols = ("num", "desc")
        self.preview_tree = ttk.Treeview(preview_frame, columns=cols, show="headings",
                                          height=8, selectmode="none")
        self.preview_tree.heading("num", text="N")
        self.preview_tree.heading("desc", text="Descricao do Servico")
        self.preview_tree.column("num", width=40, anchor=tk.CENTER)
        self.preview_tree.column("desc", width=420)

        pv_scroll = ttk.Scrollbar(preview_frame, orient=tk.VERTICAL, command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=pv_scroll.set)
        self.preview_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        pv_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # --- Botões ---
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=(10, 0))

        ttk.Button(btn_frame, text="Cancelar", command=self.dialog.destroy).pack(side=tk.RIGHT, padx=(6, 0))
        ttk.Button(btn_frame, text="Gerar OS", style="Accent.TButton",
                   command=self._gerar_os).pack(side=tk.RIGHT)

    def _on_patrimonio_select(self, event=None):
        """Quando o usuário seleciona um patrimônio, auto-preenche os dados."""
        patri = self.patrimonio_var.get().strip()
        if not patri:
            return

        for row in self.cadastro_rows:
            if row['patrimonio'] == patri:
                self.tipo_manut_var.set(row['tipo_manut'])
                self.descricao_var.set(row['descricao'])
                self.marca_var.set(row['marca'])
                self.modelo_var.set(row['modelo'])
                self.local_var.set(row['local'])
                self.unde_var.set(row['unde'])
                self._atualizar_preview()
                return

        for var in (self.tipo_manut_var, self.descricao_var, self.marca_var,
                    self.modelo_var, self.local_var, self.unde_var):
            var.set("")
        self._atualizar_preview()

    def _atualizar_preview(self, event=None):
        """Atualiza a prévia dos serviços conforme TipoManut + Periodicidade."""
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)

        tipo = self.tipo_manut_var.get().strip()
        per = self.periodicidade_var.get().strip()
        if not tipo or not per:
            return

        if per.isdigit():
            per_str = per
        else:
            per_str = str(int(float(per))) if per.replace('.', '', 1).isdigit() else per
        lookup_key = tipo + per_str

        ws_param = self.engine.workbook["ParametroManut"]

        coluna = None
        for c in range(2, (ws_param.max_column or 1) + 1):
            h = ws_param.cell(row=1, column=c).value
            if h is not None and str(h).strip() == lookup_key:
                coluna = c
                break

        if coluna is None:
            return

        idx = 1
        for r in range(2, (ws_param.max_row or 1) + 1):
            val = ws_param.cell(row=r, column=coluna).value
            if val is not None and safe_str(val).upper() == "X":
                desc = safe_str(ws_param.cell(row=r, column=1).value)
                if desc:
                    self.preview_tree.insert("", tk.END, values=(f"{idx:02d}", desc))
                    idx += 1

    def _gerar_os(self):
        """Valida dados, escreve na OSP e executa preenche_osp."""
        if not self.patrimonio_var.get().strip():
            messagebox.showwarning("Aviso", "Selecione um patrimonio.", parent=self.dialog)
            return
        if not self.periodicidade_var.get().strip():
            messagebox.showwarning("Aviso", "Informe a periodicidade.", parent=self.dialog)
            return

        if not messagebox.askyesno("Confirmar",
                                   "Gerar OS e copiar para Movimento?\n\nA planilha sera modificada e salva.",
                                   parent=self.dialog):
            return

        # --- Escreve cabeçalho na OSP ---
        ws_osp = self.engine.workbook["OSP"]

        ws_osp.cell(row=4, column=1).value = self.local_var.get().strip()      # A4 = Local
        ws_osp.cell(row=4, column=3).value = self.descricao_var.get().strip()  # C4 = Equipamento
        ws_osp.cell(row=4, column=6).value = self.marca_var.get().strip()      # F4 = Marca
        ws_osp.cell(row=4, column=8).value = self.modelo_var.get().strip()     # H4 = Modelo

        hoje = datetime.now()
        ws_osp.cell(row=6, column=1).value = self.patrimonio_var.get().strip()    # A6 = Patrimonio
        ws_osp.cell(row=6, column=2).value = self.tipo_manut_var.get().strip()    # B6 = TipoManut
        ws_osp.cell(row=6, column=3).value = self.periodicidade_var.get().strip() # C6 = Periodicidade
        ws_osp.cell(row=6, column=4).value = self.unde_var.get().strip()          # D6 = Unde
        ws_osp.cell(row=6, column=8).value = hoje                                  # H6 = Emissao

        self.parent.log(f"OS: Patrimonio={self.patrimonio_var.get().strip()}, "
                        f"Tipo={self.tipo_manut_var.get().strip()}, "
                        f"Per.={self.periodicidade_var.get().strip()}")

        # --- Executa preenche_osp ---
        self.parent.set_status("Gerando OS...")
        result = self.parent._safe_operation(
            lambda progress_callback: self.engine.preenche_osp(
                progress_callback=progress_callback,
            )
        )

        if result:
            # --- Auto-salva ---
            try:
                orig = self.engine.filepath
                if orig and orig.lower().endswith(".xlsx"):
                    save_path = orig
                else:
                    base = orig.rsplit(".", 1)[0] if orig else "planilha"
                    save_path = base + ".xlsx"
                self.engine.save(save_path)
                self.parent.log(f"Planilha salva em: {save_path}")
            except Exception as e:
                self.parent.log(f"Erro ao salvar planilha: {e}")
                self.parent.log(traceback.format_exc())
                messagebox.showwarning("Aviso",
                    f"OS gerada, mas nao foi possivel salvar:\n{e}",
                    parent=self.dialog)

            self.parent.log("OS gerada com sucesso!")
            self.parent.set_status("OS gerada com sucesso! Planilha salva.")
            self.dialog.destroy()
        else:
            self.parent.set_status("Erro ao gerar OS. Verifique o log.")


# =====================================================================
# INTERFACE GRAFICA (Tkinter)
# =====================================================================

class Application:
    """Interface grafica simplificada — apenas geracao de OS."""

    def __init__(self, root):
        self.root = root
        self.root.title("Gerador de Ordem de Servico")
        self.root.geometry("1280x760")
        self.root.minsize(900, 600)

        self.engine = MacroEngine()
        self._operating = False

        self._build_ui()

    def _build_ui(self):
        """Interface simplificada: toolbar + visualizador + log."""
        # --- Toolbar ---
        toolbar = ttk.Frame(self.root)
        toolbar.pack(fill=tk.X, padx=6, pady=4)

        ttk.Button(toolbar, text="Abrir Arquivo", command=self.open_file).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Salvar .xlsx", command=self.save_file).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Gerar OS (Assistido)", style="Accent.TButton",
                   command=self.run_gerar_os).pack(side=tk.LEFT, padx=(20, 2))

        # --- Painel dividido: visualizador + log ---
        paned = ttk.PanedWindow(self.root, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 4))

        # Visualizador de planilha
        view_frame = ttk.LabelFrame(paned, text="Visualizar Planilha")
        paned.add(view_frame, weight=3)

        top_frame = ttk.Frame(view_frame)
        top_frame.pack(fill=tk.X, padx=4, pady=4)
        ttk.Label(top_frame, text="Aba:").pack(side=tk.LEFT)
        self.sheet_combo = ttk.Combobox(top_frame, state="readonly", width=40)
        self.sheet_combo.pack(side=tk.LEFT, padx=4)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_select)

        tree_frame = ttk.Frame(view_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=(0, 4))

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        self.tree = ttk.Treeview(tree_frame, show="headings",
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        # Log
        log_frame = ttk.LabelFrame(paned, text="Log")
        paned.add(log_frame, weight=1)

        self.log_text = tk.Text(log_frame, height=6, wrap=tk.WORD,
                                 state=tk.DISABLED, font=("Consolas", 9))
        log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # --- Barra de status ---
        self.status_var = tk.StringVar(value="Pronto. Abra um arquivo para comecar.")
        status_bar = ttk.Label(self.root, textvariable=self.status_var,
                               relief=tk.SUNKEN, anchor=tk.W, padding=(6, 2))
        status_bar.pack(fill=tk.X)

    # ------------------------------------------------------------------
    # LOG / STATUS
    # ------------------------------------------------------------------

    def log(self, msg, tag=""):
        """Adiciona mensagem ao log."""
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} {msg}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)
        self.root.update_idletasks()

    def set_status(self, msg):
        """Atualiza barra de status."""
        self.status_var.set(msg)
        self.root.update_idletasks()

    # ------------------------------------------------------------------
    # ARQUIVO
    # ------------------------------------------------------------------

    def open_file(self):
        """Abre arquivo .xls ou .xlsx."""
        filepath = filedialog.askopenfilename(
            title="Selecionar planilha",
            filetypes=[
                ("Arquivos Excel", "*.xls *.xlsx"),
                ("Excel 97-2003", "*.xls"),
                ("Excel Open XML", "*.xlsx"),
            ]
        )
        if not filepath:
            return

        self.set_status("Carregando arquivo...")
        self.log(f"Abrindo arquivo: {os.path.basename(filepath)}")

        try:
            if filepath.lower().endswith(".xlsx"):
                self.engine.load_xlsx(filepath)
            else:
                self.engine.load_xls(filepath)

            sheets = self.engine.get_sheet_names()
            self.sheet_combo["values"] = sheets
            if sheets:
                self.sheet_combo.current(0)
                self.on_sheet_select()

            self.set_status(f"{os.path.basename(filepath)} — {len(sheets)} abas")
            self.log(f"Arquivo carregado: {len(sheets)} abas encontradas.")
        except Exception as e:
            self.log(f"ERRO ao abrir arquivo: {e}")
            self.log(traceback.format_exc())
            messagebox.showerror("Erro", f"Nao foi possivel abrir o arquivo:\n{e}")

    def save_file(self):
        """Salva workbook como .xlsx."""
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel .xlsx", "*.xlsx")]
        )
        if filepath:
            try:
                self.engine.save(filepath)
                self.log(f"Arquivo salvo: {filepath}")
                self.set_status(f"Salvo: {os.path.basename(filepath)}")
            except Exception as e:
                self.log(f"ERRO ao salvar: {e}")
                messagebox.showerror("Erro", f"Nao foi possivel salvar:\n{e}")

    # ------------------------------------------------------------------
    # VISUALIZACAO
    # ------------------------------------------------------------------

    def on_sheet_select(self, event=None):
        """Atualiza visualizacao ao selecionar aba."""
        name = self.sheet_combo.get()
        if not name:
            return
        data = self.engine.get_sheet_data(name)
        self._populate_tree(data)

    def _populate_tree(self, data):
        """Preenche a TreeView com os dados."""
        self.tree.delete(*self.tree.get_children())
        if not data:
            return

        cols = [f"Col {i+1}" for i in range(len(data[0]))]
        self.tree["columns"] = cols
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, minwidth=60)

        for row in data:
            values = [str(v) if v is not None else "" for v in row]
            self.tree.insert("", tk.END, values=values)

    def refresh_view(self):
        """Recarrega a aba atual."""
        self.on_sheet_select()

    # ------------------------------------------------------------------
    # OPERACAO SEGURA
    # ------------------------------------------------------------------

    def _safe_operation(self, func):
        """Executa operacao com protecao de concorrencia."""
        if self._operating:
            messagebox.showinfo("Aguarde", "Outra operacao esta em execucao.")
            return
        self._operating = True
        try:
            self.root.config(cursor="watch")
            self.root.update_idletasks()

            def progress(msg):
                self.log(msg)
                self.root.update_idletasks()

            result = func(progress_callback=progress)
            return result
        except Exception as e:
            self.log(f"ERRO: {e}")
            self.log(traceback.format_exc())
            messagebox.showerror("Erro", str(e))
            return None
        finally:
            self.root.config(cursor="")
            self._operating = False
            self.refresh_view()

    # ------------------------------------------------------------------
    # GERAR OS
    # ------------------------------------------------------------------

    def run_gerar_os(self):
        """Handler: abre o dialogo assistido Gerar OS."""
        if not self.engine.workbook:
            messagebox.showwarning("Aviso", "Abra um arquivo primeiro.")
            return
        try:
            GerarOSDialog(self, self.engine)
        except Exception as e:
            self.log(f"ERRO ao abrir dialogo Gerar OS: {e}")
            self.log(traceback.format_exc())
            messagebox.showerror("Erro", str(e))


# =====================================================================
# MAIN
# =====================================================================

if __name__ == "__main__":
    root = tk.Tk()

    style = ttk.Style()
    available_themes = style.theme_names()
    for preferred in ("vista", "clam", "alt", "default"):
        if preferred in available_themes:
            style.theme_use(preferred)
            break

    style.configure("Accent.TButton",
                    font=("", 9, "bold"),
                    foreground="#ffffff",
                    background="#2c5f8a")

    app = Application(root)
    root.mainloop()
